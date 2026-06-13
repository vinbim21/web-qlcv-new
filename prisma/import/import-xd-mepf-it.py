# -*- coding: utf-8 -*-
"""Nạp Tiến độ thực tế (cột 'Thực tế hoàn thành') + Timesheet hằng ngày
từ WM_New (1).xlsx (sheet XD/MEPF/IT) vào DB local.

Quy tắc (đã chốt với người dùng):
  - Timesheet: THAY THẾ theo khoảng ngày (xoá entry của các user trong [min,max] rồi nạp lại).
  - Dòng giờ KHÔNG có mã việc -> taskId = null (vẫn ghi nhận giờ công).
  - Cột 'Thực tế hoàn thành' có ngày -> actualEnd = ngày, status = HOAN_THANH, progress = 100. Nhiều ngày/việc -> lấy MAX.
  - Bỏ qua mã không khớp + giá trị không phải ngày, có log.

Dùng:
  python prisma/import/import-xd-mepf-it.py dry   # chỉ in báo cáo
  python prisma/import/import-xd-mepf-it.py sql    # sinh prisma/import/_load-xd.sql
Cần trước: _map_users.tsv (id<TAB>fullName), _map_tasks.tsv (id<TAB>sumId<TAB>projectId).
"""
import sys, os, warnings
from datetime import datetime, date
warnings.filterwarnings("ignore")
import openpyxl

ROOT = os.getcwd()
XLSX = r"C:\Users\toanpc\Downloads\WM_New (1).xlsx"
HERE = os.path.join(ROOT, "prisma", "import")
SHEETS = ["XD", "MEPF", "IT"]

def s(v):
    if v is None: return ""
    if isinstance(v, (datetime, date)): return v.isoformat()[:10]
    return str(v).strip()
def isdate(v): return isinstance(v, (datetime, date))
def asdate(v): return v.date() if isinstance(v, datetime) else v
def num(v): return float(v) if isinstance(v, (int, float)) else None
def q(t): return "'" + t.replace("'", "''") + "'"

def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "dry"
    # maps
    user_by_name = {}
    for line in open(os.path.join(HERE, "_map_users.tsv"), encoding="utf-8"):
        line = line.rstrip("\n")
        if not line: continue
        uid, name = (line.split("\t") + ["", ""])[:2]
        user_by_name[name.strip()] = uid
    task_by_sum = {}
    proj_by_task = {}
    for line in open(os.path.join(HERE, "_map_tasks.tsv"), encoding="utf-8"):
        line = line.rstrip("\n")
        if not line: continue
        tid, sumid, pid = (line.split("\t") + ["", "", ""])[:3]
        task_by_sum[sumid.strip()] = tid
        proj_by_task[tid] = pid.strip()

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

    ts = {}             # (userId, taskId|None, dateISO) -> [hours, notes set]
    actual = {}         # taskId -> max date
    skip_unmatched_id = set()
    skip_bad_actual = []       # (sumId, raw)
    skip_actual_nomatch = set()
    no_id_hours = 0.0
    summary_rows = 0
    dmin = dmax = None
    used_user_ids = set()

    for name in SHEETS:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        for r in ws.iter_rows(min_row=4, values_only=True):
            person = s(r[0]) if len(r) > 0 else ""
            sumid  = s(r[1]) if len(r) > 1 else ""
            c8     = r[7] if len(r) > 7 else None
            if not person or person.lower() == "all":
                continue
            uid = user_by_name.get(person)
            if not uid:
                continue  # 14/14 khớp nên không xảy ra
            tid = task_by_sum.get(sumid) if sumid else None
            if sumid and not tid:
                skip_unmatched_id.add(sumid)

            # Nhãn cấp (L2..L5) — để nhận diện dòng tổng & làm note cho việc chung
            level_parts = [s(r[i]) for i in (3, 4, 5, 6) if len(r) > i]
            label = " / ".join(p for p in level_parts if p and p.lower() != "all")
            has_level = bool(label)
            # Dòng TỔNG theo người: không mã việc & không cấp -> BỎ (tránh nhân đôi giờ)
            is_summary = (not sumid) and (not has_level)
            if is_summary:
                summary_rows += 1

            # --- Thực tế hoàn thành ---
            if c8 not in (None, ""):
                if isdate(c8):
                    if tid:
                        d = asdate(c8).isoformat()
                        if tid not in actual or d > actual[tid]:
                            actual[tid] = d
                    else:
                        skip_actual_nomatch.add(sumid or "(trống)")
                else:
                    skip_bad_actual.append((sumid, s(c8)))

            # --- Timesheet (gap-aware) ---
            if is_summary:
                continue  # bỏ qua hoàn toàn dòng tổng
            c = 9  # 0-indexed cột 10
            while c < len(r):
                v = r[c]
                if isdate(v):
                    h = num(r[c+2]) if c+2 < len(r) else None
                    note = s(r[c+1]) if c+1 < len(r) else ""
                    if not note and not tid:
                        note = label  # việc chung không mã -> ghi nhãn cấp làm note
                    if h and h > 0:
                        d = asdate(v).isoformat()
                        key = (uid, tid, d)
                        if key not in ts: ts[key] = [0.0, set()]
                        ts[key][0] += h
                        if note: ts[key][1].add(note)
                        used_user_ids.add(uid)
                        if not tid: no_id_hours += h
                        dd = asdate(v)
                        dmin = dd if dmin is None or dd < dmin else dmin
                        dmax = dd if dmax is None or dd > dmax else dmax
                    c += 3
                else:
                    c += 1

    total_hours = sum(v for v, _ in ts.values())
    n_noid = sum(1 for (_, t, _) in ts if t is None)

    print("================ DRY-RUN REPORT ================")
    print("Khoảng ngày timesheet : %s -> %s" % (dmin, dmax))
    print("Số dòng timesheet (gộp theo user+việc+ngày): %d" % len(ts))
    print("  - có mã việc : %d" % (len(ts) - n_noid))
    print("  - KHÔNG mã việc (taskId=null): %d  (tổng %.2f h)" % (n_noid, no_id_hours))
    print("Dòng TỔNG/người đã BỎ (chống nhân đôi): %d dòng" % summary_rows)
    print("Tổng giờ công SẼ NẠP : %.2f h" % total_hours)
    print("Số user liên quan    : %d" % len(used_user_ids))
    print("Task cập nhật HOÀN THÀNH (actualEnd+100%%): %d" % len(actual))
    print("------- BỎ QUA (báo để xem tay) -------")
    print("Mã việc không khớp DB: %d  %s" % (len(skip_unmatched_id), sorted(skip_unmatched_id)))
    print("'Thực tế HT' không phải ngày: %s" % skip_bad_actual)
    print("'Thực tế HT' nhưng việc không khớp: %s" % sorted(skip_actual_nomatch))

    if mode != "sql":
        return

    # --- sinh SQL ---
    out = os.path.join(HERE, "_load-xd.sql")
    lines = []
    lines.append("BEGIN;")
    lines.append("SET LOCAL search_path = public;")
    # Phần A: actualEnd + HOAN_THANH + 100
    for tid, d in sorted(actual.items()):
        lines.append(
            'UPDATE "Task" SET "actualEnd"=%s::timestamp, "status"=\'HOAN_THANH\', '
            '"progressPercent"=100, "updatedAt"=now() WHERE id=%s;' % (q(d), q(tid))
        )
    # Phần B: thay thế timesheet theo khoảng ngày cho các user liên quan
    uids = sorted(used_user_ids)
    if uids and dmin and dmax:
        in_list = ",".join(q(u) for u in uids)
        lines.append(
            'DELETE FROM "TimeSheetEntry" WHERE "userId" IN (%s) '
            'AND date BETWEEN %s::date AND %s::date;'
            % (in_list, q(dmin.isoformat()), q(dmax.isoformat()))
        )
    i = 0
    for (uid, tid, d), (h, notes) in ts.items():
        i += 1
        eid = "tsxls_%05d" % i
        note = " | ".join(sorted(notes)) if notes else None
        pid = proj_by_task.get(tid) if tid else ""
        cols = '"id","userId","taskId","projectId","date","hours","note","createdAt","updatedAt"'
        vals = "%s,%s,%s,%s,%s::date,%s,%s,now(),now()" % (
            q(eid), q(uid),
            q(tid) if tid else "NULL",
            q(pid) if pid else "NULL",
            q(d), ("%.2f" % h),
            q(note) if note else "NULL",
        )
        lines.append('INSERT INTO "TimeSheetEntry" (%s) VALUES (%s);' % (cols, vals))
    lines.append("COMMIT;")
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print("\nSQL -> %s  (%d câu lệnh)" % (out, len(lines)))

if __name__ == "__main__":
    main()
