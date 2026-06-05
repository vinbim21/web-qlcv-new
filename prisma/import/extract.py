# -*- coding: utf-8 -*-
"""Trích xuất WM_New.xlsx -> prisma/import/data.json (cho run-all.ts nạp vào DB).
Chạy: python prisma/import/extract.py
"""
import json, os, sys, warnings
from datetime import datetime, date
warnings.filterwarnings("ignore")
import openpyxl

ROOT = os.getcwd()
XLSX = os.path.join(ROOT, "WM_New.xlsx")
OUT = os.path.join(ROOT, "prisma", "import", "data.json")

# Cấu hình cột (1-indexed) mỗi Bảng giao việc. Dữ liệu từ dòng 3.
TASK_SHEETS = [
    {"sheet": "Bảng 1", "wg": "1", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":8,"phase":None,"assignees":[9,10,11],"start":12,"end":13},
    {"sheet": "Bảng 2", "wg": "2", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":8,"phase":None,"assignees":[9,10,11],"start":12,"end":13},
    {"sheet": "Bảng 3", "wg": "3", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":9,"phase":8,"assignees":[10,11,12,13],"start":14,"end":15},
    {"sheet": "Bảng 4", "wg": "4", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":8,"phase":None,"assignees":[9,10,11],"start":12,"end":13},
    {"sheet": "Bảng 5", "wg": "5", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":8,"phase":None,"assignees":[9,10,11],"start":12,"end":13},
    {"sheet": "Bảng 6", "wg": "6", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":8,"phase":None,"assignees":[9],"start":10,"end":11},
    {"sheet": "Bảng 7", "wg": "7", "sumId":1,"subId":2,"l2":4,"l3":5,"l4":6,"l5":7,"priority":8,"phase":None,"assignees":[9,10,11],"start":12,"end":13},
]
TS_SHEETS = ["XD", "MEPF", "IT"]

# Sheet "Data": cột Level 2/3/5 theo từng nhóm (1-indexed). data từ dòng 4.
# (wg, {2: colL2, 3: colL3, 5: colL5})  — nhóm nào thiếu cấp thì bỏ qua.
DATA_COLS = [
    ("1", {2: 2, 3: 3, 5: 4}),     # Xây dựng HTTC BIM
    ("2", {2: 5, 3: 6, 5: 7}),     # Đào tạo BIM
    ("3", {2: 8, 3: 9, 5: 11}),    # Quản lý BIM (cột 10=Quy mô, bỏ)
    ("4", {5: 12}),                # Thanh tra BIM (chỉ Level 5)
    ("5", {2: 13, 3: 14, 5: 15}),  # Phát triển BIM Tools
    ("6", {2: 16, 3: 17, 5: 18}),  # Quản lý phần mềm
    ("7", {2: 19, 3: 20, 5: 21}),  # Công việc khác
]

def sval(v):
    if v is None: return ""
    if isinstance(v, (datetime, date)): return v.isoformat()[:10]
    return str(v).strip()

def dval(v):
    if isinstance(v, (datetime, date)): return v.isoformat()[:10]
    return None

def nval(v):
    if isinstance(v, (int, float)): return float(v)
    return None

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    tasks = []
    for cfg in TASK_SHEETS:
        if cfg["sheet"] not in wb.sheetnames: continue
        ws = wb[cfg["sheet"]]
        for row in ws.iter_rows(min_row=3):
            def cell(c):
                idx = c - 1
                return row[idx].value if idx < len(row) else None
            sumId = sval(cell(cfg["sumId"]))
            l5 = sval(cell(cfg["l5"])); l3 = sval(cell(cfg["l3"])); l2 = sval(cell(cfg["l2"]))
            if not (sumId or l5 or l3 or l2): continue
            assignees = []
            for col in cfg["assignees"]:
                raw = sval(cell(col))
                for part in raw.replace("/", ",").replace(";", ",").split(","):
                    p = part.strip()
                    if p and p.lower() != "all" and p not in assignees:
                        assignees.append(p)
            tasks.append({
                "wg": cfg["wg"], "sumId": sumId, "subId": sval(cell(cfg["subId"])),
                "l2": l2, "l3": l3, "l4": sval(cell(cfg["l4"])), "l5": l5,
                "priority": sval(cell(cfg["priority"])),
                "phase": sval(cell(cfg["phase"])) if cfg["phase"] else "",
                "assignees": assignees,
                "start": dval(cell(cfg["start"])), "end": dval(cell(cfg["end"])),
            })

    timesheets = []
    for name in TS_SHEETS:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        for row in ws.iter_rows(min_row=4):
            def cell(c):
                idx = c - 1
                return row[idx].value if idx < len(row) else None
            person = sval(cell(1))
            if not person or person.lower() == "all": continue
            taskSum = sval(cell(2))
            c = 10
            ncol = len(row)
            while c + 2 <= ncol:
                d = dval(cell(c)); content = sval(cell(c+1)); hours = nval(cell(c+2))
                if d and hours and hours > 0:
                    timesheets.append({"person": person, "taskSum": taskSum, "date": d, "hours": hours, "note": content})
                c += 3

    # Danh mục Level 2/3/5 từ sheet "Data"
    catalog = []
    if "Data" in wb.sheetnames:
        ws = wb["Data"]
        seen = set()
        for row in ws.iter_rows(min_row=4):
            for wg, cols in DATA_COLS:
                for level, col in cols.items():
                    idx = col - 1
                    val = sval(row[idx].value) if idx < len(row) else ""
                    if val and (wg, level, val) not in seen:
                        seen.add((wg, level, val))
                        catalog.append({"wg": wg, "level": level, "value": val})

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"tasks": tasks, "timesheets": timesheets, "catalog": catalog}, f, ensure_ascii=False)
    print("tasks=%d timesheets=%d catalog=%d -> %s" % (len(tasks), len(timesheets), len(catalog), OUT))

if __name__ == "__main__":
    main()
